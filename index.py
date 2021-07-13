from openpyxl import Workbook,load_workbook
import time
import logging
from concurrent.futures import ThreadPoolExecutor
import threading
from functions import explorar


logging.basicConfig(level=logging.DEBUG, format='%(threadName)s: %(message)s')


#directorios
# directoryraiz = "C:\Prueba"
# directoryraiz = "C:\Prueba"
# directoryraiz_old = "F:\CORALES BASE DE DATOS\juanCorales"
# directoryraiz = 'C:\\Users\\usrlabsis30\Documents\\BACKUPCORALES\\juanCorales'
# directoryraiz = 'C:\\Users\\usrlabsis30\\Documents\\BACKUPCORALES\\RAUL'
directoryraiz = 'C:\\Users\\usrlabsis30\\Documents\\BACKUPCORALES\\andres'
# directoryraiz = 'C:\\Users\\usrlabsis30\\Documents\BACKUPCORALES\\BASE DE DATOS DE PASTOS MARINOS'


#funcion
def ReadAllRow(listDirectory):

    nombre_hilo = threading.current_thread().name
    print("Ejecutando hilo :",nombre_hilo)
    # newbook = Workbook()
    # sheet = newbook.active

 #GUARDAR
    if nombre_hilo == 'ThreadPoolExecutor-0_0':
        newbook = load_workbook("C:\Prueba\parte1.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_1':
        newbook = load_workbook("C:\Prueba\parte2.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_2':
        newbook = load_workbook("C:\Prueba\parte3.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_3':
        newbook = load_workbook("C:\Prueba\parte4.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_4':
        newbook = load_workbook("C:\Prueba\parte5.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_5':
        newbook = load_workbook("C:\Prueba\parte6.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_6':
        newbook = load_workbook("C:\Prueba\parte7.xlsx")
        sheet = newbook['Sheet']
    elif nombre_hilo == 'ThreadPoolExecutor-0_7':
        newbook = load_workbook("C:\Prueba\parte8.xlsx")
        sheet = newbook['Sheet']
    else:
        newbook.save("C:\Prueba\otros.xlsx") 

    sinHojaDatos = 0
    con = 1
    for directory in listDirectory:
        list = []
        try:
            print("FILE",con," PID :",nombre_hilo," -> DIREC :",directory)
            wb = load_workbook(filename=directory, read_only=False ,data_only=True)
            # , keep_vba=True,
            ws = wb['DATOS']
            if ws:
                list.append(directory)  
                for row in ws.iter_rows(10, 10):
                    for cell in row:
                        list.append(cell.value)
                #AGREGAR
                sheet.append(list)
              
        except KeyError:
            print("Hoja '{}' No encontrada en el libro de la ruta: '{}'".format('DATOS', directory))
            sinHojaDatos +=1
        con += 1    
    
    #GUARDAR
    if nombre_hilo == 'ThreadPoolExecutor-0_0':
        newbook.save("C:\Prueba\parte1.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_1':
        newbook.save("C:\Prueba\parte2.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_2':
        newbook.save("C:\Prueba\parte3.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_3':
        newbook.save("C:\Prueba\parte4.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_4':
        newbook.save("C:\Prueba\parte5.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_5':
        newbook.save("C:\Prueba\parte6.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_6':
        newbook.save("C:\Prueba\parte7.xlsx")
    elif nombre_hilo == 'ThreadPoolExecutor-0_7':
        newbook.save("C:\Prueba\parte8.xlsx")
    else:
        newbook.save("C:\Prueba\otros.xlsx") 

    print("==========================================")
    print(f"* hilo :{nombre_hilo} OK!")
    print(f"* Plantillas sin hoja de Datos :{sinHojaDatos}")
    print(f"* Plantillas con hoja de Datos :{len(listDirectory)-sinHojaDatos}")
    print("==========================================")

    # logging.info(f'Termino')



#correr procesos
if __name__ == '__main__':
    
    #config numero de nucleos
    executor = ThreadPoolExecutor(max_workers=8)

    # explorar todo el directorio
    listDirectory = explorar(directoryraiz)
    # print(directoryraiz)
    # print("===============================")
    # print(listDirectory)
    
    # dividiendo listDirectory
    x = 30
    final_list= lambda listDirectory, x: [listDirectory[i:i+x] for i in range(0, len(listDirectory), x)]
    output=final_list(listDirectory, x)

    print("Total archivos .xlsm :",len(listDirectory),"/ 8 procesadores")
    for  index, ou in enumerate(output):
        print(f"Tamaño parte {index+1} :[", len(output[index]),"] Registros")


    for  index, ou in enumerate(output):
        print("===============================")
        executor.submit(ReadAllRow,output[index])
        print("===============================")
    

