from openpyxl import Workbook,load_workbook
import time
import logging
from concurrent.futures import ThreadPoolExecutor
import threading
from functions import explorar
from pathlib import Path


logging.basicConfig(level=logging.DEBUG, format='%(threadName)s: %(message)s')

MAX_WORKERS = 8
CWD = Path.cwd()

directoryraiz = CWD.parent/'input'


#funcion
def ReadAllRow(listDirectory):
    start = time.time()
    nombre_hilo = threading.current_thread().name
    print(f"Ejecutando hilo : {nombre_hilo}")
 #GUARDAR
    loader_workbook = dict({
        f'ThreadPoolExecutor-0_{i}': CWD.parent/f"output/parte{i+1}.xlsx" for i in range(MAX_WORKERS)})

    if nombre_hilo in loader_workbook:
        newbook = load_workbook(loader_workbook[nombre_hilo])
        sheet = newbook['Sheet']
    else:
        newbook.save(CWD.parent/"output/otros.xlsx") 

    sinHojaDatos = 0
    con = 1
    for directory in listDirectory:
        list = []
        try:
            print(f"FILE {con} PID :{nombre_hilo} -> DIREC :{directory}")
            wb = load_workbook(filename=directory, read_only=False ,data_only=True)
            # , keep_vba=True,
            ws = wb['DATOS']
            if ws:
                list.append(directory)  
                for row in ws.iter_rows(10, 10):
                    for cell in row:
                        list.append(cell.value)
                #AGREGAR
                sheet.append(list)
              
        except KeyError:
            print("Hoja '{}' No encontrada en el libro de la ruta: '{}'".format('DATOS', directory))
            sinHojaDatos +=1
        con += 1    
    
    #GUARDAR
    if nombre_hilo in loader_workbook:
        newbook.save(loader_workbook[nombre_hilo])
    else:
        newbook.save(CWD.parent/"output/otros.xlsx") 

    print(f"""
    ==========================================
    * hilo :{nombre_hilo} OK!
    * Plantillas sin hoja de Datos : {sinHojaDatos}
    * Plantillas con hoja de Datos : {len(listDirectory)-sinHojaDatos}
    * Took: {time.time()-start:.4f} secs
    ==========================================""")



#correr procesos
if __name__ == '__main__':
    
    #config numero de nucleos
    executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)

    # explorar todo el directorio
    listDirectory = explorar(directoryraiz)
    
    # dividiendo listDirectory
    x = 30
    final_list= lambda listDirectory, x: [listDirectory[i:i+x] for i in range(0, len(listDirectory), x)]
    output=final_list(listDirectory, x)

    print(f"Total archivos .xlsm :{len(listDirectory)}/ {MAX_WORKERS} procesadores")
    for  index, ou in enumerate(output):
        print(f"Tamaño parte {index+1} :[{len(output[index])}] Registros")


    for  index, ou in enumerate(output):
        print(f"""
        ===============================
        {executor.submit(ReadAllRow,output[index])}
        ===============================""")


