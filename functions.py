import os
# from os.path import isfile, join, isdir
# from functionsXlsx import *
import multiprocessing as mp
from openpyxl import load_workbook

LOCK = mp.Lock()

def ReadAllRow(directory, q):
    directoryout = "C:\Prueba\OUT.xlsx"
    list = []
    sheetname = 'DATOS'
    # sheet = book.worksheets[0]

    # book = load_workbook(directoryout)
    # sheet = book['Sheet']


    try:
        wb = load_workbook(filename=directory, read_only=False ,data_only=True)
        # , keep_vba=True,
        ws = wb[sheetname]
        print("DIREC ",directory)
        if ws:
            for row in ws.iter_rows(10, 10):
                for cell in row:
                    list.append(cell.value)
            list.append(directory)  
            # sheet.append(list)
           
    except KeyError:
        print("Worksheet '{}' not found for workbook '{}'.Adding...".format(sheetname, directory))
    # with LOCK:
    # book.save(directoryout)
    q.put(list) 
    return list

def explorar(directory):
    listaDir = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsm'):
                listaDir.append(os.path.join(root, file))
    # print(listaDir)            
    return listaDir

    
fn = 'c:/temp/temp.txt'
