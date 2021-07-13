from openpyxl import Workbook, load_workbook
import xlwings as xw


def crearXlsx(directory):
    book = Workbook()
    # book.create_sheet("DATOS", 0)
    sheet = book.active
    return sheet


def defineHeaderXlsx(directory, colunmName):
    wb = load_workbook(directory)
    hoja = wb['DATOS']
    con = 1
    for i in colunmName:
        hoja.cell(row=1, column=con).value = i
        con += 1
    wb.save(directory)


# leer
def ReadXlsx(directory):
    wb = load_workbook(directory)
    hoja = wb['DATOS']
    x = hoja['E5']
    print(x.value)


def addData(directory, registros, full_file_paths, colunmName):

    wb = load_workbook(directory)
    hoja = wb['DATOS']

    con = 2
    for i in full_file_paths:
        print(i)
        hoja['F{}'.format(con)] = i
        con += 1

    wb.save(directory)

    rows = 2
    for reg in registros:

        hoja.cell(row=rows, column=1).value = reg[0].value
        hoja.cell(row=rows, column=2).value = reg[1].value
        hoja.cell(row=rows, column=3).value = reg[2].value
        hoja.cell(row=rows, column=4).value = reg[3].value
        hoja.cell(row=rows, column=5).value = reg[4].value
        rows += 1

    wb.save(directory)


def loadXlsx(directory):
    wb = load_workbook(directory)
    #hoja = wb['Datos']

    # cambia nombre hoja
    wb.create_sheet("DATOS", 1)
    wb.save(directory)


# escribir
def writeXlsx(directory):
    wb = load_workbook(directory)
    hoja = wb['DATOS']
    hoja['A2'] = 'jeison diaz palmera'
    wb.save(directory)

# leer


def ReadXlsx(directory):
    wb = load_workbook(filename=directory, read_only=False,
                       keep_vba=True, data_only=True)
    hoja = wb['DATOS']
    return [hoja['D5'], hoja['E5'], hoja['N5'], hoja['X5'], hoja['AA5']]


# leer rango
def ReadRangeXlsx(directory):
    wb = load_workbook(directory)
    hoja = wb['DATOS']
    return hoja['A1':'C1']
